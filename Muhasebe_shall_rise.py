import openpyxl as xls
import tkinter as tk
import tkinter.filedialog
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showinfo
from tkinter.messagebox import showinfo


master = tk.Tk()
master.geometry("200x125")
master.title("Muhasebe Shall Rise")

def Select_File():
    global Path #we are doing our Path global so we can use it on another function
    Path = tk.StringVar() #Storing it as a string
    Path = tkinter.filedialog.askopenfilename(initialdir="/", filetypes=(
    ("excel files", "*.xlsx *.xls"), ("all files", "*.*"))) #we are getting our path of our excel file
    print(Path)
    return Path

Select_File_Button = tk.Button(master,text="Select File",command=lambda: Select_File())
Select_File_Button.grid(padx=65,pady=20,row=1)

Calculate_Button = tk.Button(master,text="RUN",command=lambda: Calculate())
Calculate_Button.grid()

def Calculate():
    wb = xls.load_workbook(Path)
    ws = wb.active
    ws.insert_rows(1,4)
    for row in range(6,ws.max_row+1):
        price = (ws.cell(row=row, column=26).value)  # Satış Fiyatını depolayan variable

        #Sipariş numaralarını depolayan variable
        first = ws.cell(row=row,column=8).value
        m_second = ws.cell(row=row-1,column=8).value
        m_third = ws.cell(row=row-2,column=8).value
        m_forth = ws.cell(row=row-3,column=8).value
        m_fifth = ws.cell(row=row - 4, column=8).value

        brand = ws.cell(row=row,column=19).value  #Markayı kontrol eden variable
        brand_2 = ws.cell(row=row-1,column=19).value
        brand_3 = ws.cell(row=row-2,column=19).value
        brand_4 = ws.cell(row=row-3, column=19).value
        brand_5 = ws.cell(row=row - 4, column=19).value

        company= ws.cell(row=row,column=3).value #Kargo şirketini kontrol eden variable
        print(type(float(price)))
        if price == None:
            continue

        elif first == m_second == m_third == m_forth == m_fifth and brand == brand_2 == brand_3 == brand_4 == brand_5  and company == "Sürat Kargo Marketplace": #4 lü sipariş Sürat
            ws.cell(row=row,column=27,value=12)
            ws.cell(row=row-1,column=27,value=12)
            ws.cell(row=row-2, column=27, value=12)
            ws.cell(row=row-3,column=27,value=12)
            ws.cell(row=row - 4, column=27, value=12)

        elif first == m_second == m_third == m_forth and brand == brand_2 == brand_3 == brand_4  and company == "Sürat Kargo Marketplace": #4 lü sipariş Sürat
            ws.cell(row=row,column=27,value=15)
            ws.cell(row=row-1,column=27,value=15)
            ws.cell(row=row-2, column=27, value=15)
            ws.cell(row=row-3,column=27,value=15)

        elif first == m_second == m_third and brand == brand_2 == brand_3 and company == "Sürat Kargo Marketplace": #3 lü sipariş Sürat
            ws.cell(row=row,column=27,value=20)
            ws.cell(row=row-1,column=27,value=20)
            ws.cell(row=row-2, column=27, value=20)

        elif first == m_second and brand == brand_2 and company == "Sürat Kargo Marketplace": #2 lü sipariş Sürat
            ws.cell(row=row,column=27,value=30)
            ws.cell(row=row-1,column=27,value=30)

        elif first == m_second == m_third == m_forth and brand == brand_2 == brand_3 == brand_4 and company == "MNG Kargo Marketplace": #4 lü sipariş MNG
            ws.cell(row=row,column=27,value=15)
            ws.cell(row=row-1,column=27,value=15)
            ws.cell(row=row-2, column=27, value=15)
            ws.cell(row=row-3, column=27, value=15)

        elif first == m_second == m_third and brand == brand_2 == brand_3 and company == "MNG Kargo Marketplace":  #3lü sipariş MNG
            ws.cell(row=row,column=27,value=20)
            ws.cell(row=row-1,column=27,value=20)
            ws.cell(row=row-2, column=27, value=20)

        elif first == m_second and brand == brand_2 and company == "MNG Kargo Marketplace": #2 li sipariş MNG
            ws.cell(row=row,column=27,value=30)
            ws.cell(row=row-1,column=27,value=30)

        elif float(price) <= 200: #Sipariş fiyatı 200 den az ise
            ws.cell(row=row,column=27,value=26.491)
        elif float(price) >= 200 and company == "Sürat Kargo Marketplace": #Sipariş fiyatı 75 ten çok ve Sürat ise
            ws.cell(row=row,column=27,value=60)
        elif float(price) >= 200 and company == "MNG Kargo Marketplace": #Sipariş fiyatı 75 ten çok ve MNG ise
            ws.cell(row=row,column=27,value=60)
        elif isinstance(price,str) == True:
            continue
        else:
            print("An Error Occured")
    tk.messagebox.showinfo("Done","Process is Done")
    ws.delete_rows(1,4)
    wb.save(Path)
master.mainloop()
