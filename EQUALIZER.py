import openpyxl as xls
import tkinter as tk
import tkinter.filedialog

master = tk.Tk()
master.geometry("200x125")
master.title("Muhasebe Shall Rise")

def Select_File():
    global Path #we are doing our Path global so we can use it on another function
    Path = tk.StringVar() #Storing it as a string
    Path = tkinter.filedialog.askopenfilename(initialdir="/", filetypes=(
    ("excel files", "*.xlsx *.xls"), ("all files", "*.*"))) #we are getting our path of our excel file
    print(Path)
    return Path

def Select_Supplier():
    global Path2 #we are doing our Path global so we can use it on another function
    Path2 = tk.StringVar() #Storing it as a string
    Path2 = tkinter.filedialog.askopenfilename(initialdir="/", filetypes=(
    ("excel files", "*.xlsx *.xls"), ("all files", "*.*"))) #we are getting our path of our excel file
    print(Path2)
    return Path2

Select_File_Button_2 = tk.Button(master,text="Supplier",command=lambda: Select_Supplier())
Select_File_Button_2.grid(padx=50,pady=5,row=1)

Select_File_Button = tk.Button(master,text="Trendyol",command=lambda: Select_File())
Select_File_Button.grid(padx=65,pady=5,row=2)

Equalizer_Button = tk.Button(master,text="EQUALIZER",command=lambda: Equalizer())
Equalizer_Button.grid()




def Equalizer():
    wb = xls.load_workbook(Path)
    ws = wb.active
    wb1 = xls.load_workbook(Path2)
    ws1 = wb1.active
    t_kargo_list = []
    t_kargo_dict = {}
    bn_kargo_dict = {}
    for i in range(2,ws.max_row+1): # Trendyol Kargo Kodu : Kaç Defa Geçti Dictionary
        t_kargo_kodu = (ws.cell(row=i,column=7).value)
        t_kargo_list.append(t_kargo_kodu)
        count = t_kargo_list.count(t_kargo_kodu)
        t_kargo_dict[t_kargo_kodu] = count

    for bn in range(2,ws1.max_row+1): # Tedarikçi Kargo Kodu : Ürün Alış bedeli Dictionary
        bn_kargo_kodu = ws1.cell(row=bn, column=56).value
        price = ws1.cell(row=bn, column=13).value
        bn_kargo_dict[bn_kargo_kodu] = price

    for i in range(2,ws.max_row+1): # Tedarikçi ve Trendyol Kargo Kodları denk geldiğinde bedel/sayım yapan kod (düzelt)
        k_kargo = ws.cell(row=i,column=7).value
        for j in bn_kargo_dict:
            for t in t_kargo_dict:
                if k_kargo == j == t:
                    value = bn_kargo_dict[j]/t_kargo_dict[t]
                    ws.cell(row=i,column=34,value=value) #Kod sonraki ay çalışmazsa column'ı 33'ten 32 ye çek.

    wb.save(Path)

master.mainloop()